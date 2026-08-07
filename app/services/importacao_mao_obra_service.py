import csv
import io
import logging
import re
import unicodedata

from datetime import datetime
from decimal import Decimal, InvalidOperation, ROUND_HALF_UP
from pathlib import Path
from typing import Any
from zoneinfo import ZoneInfo

from fastapi import Depends
from openpyxl import load_workbook

from app.clients.sankhya_client import (
    SankhyaClient,
    get_sankhya_client,
)
from app.schemas.importacao_mao_obra import (
    MaoObraImportacaoResponse,
    MaoObraLinhaPreview,
    MaoObraPreviewResponse,
    MaoObraResultadoLinha,
)


logger = logging.getLogger(__name__)


class ImportacaoMaoObraService:
    """
    Importação de valores de mão de obra para o Sankhya.

    Regras:
    - aceita XLSX e CSV;
    - para XLSX, lê exclusivamente a aba OBRAS;
    - identifica as colunas Cód e Valor Acumulado Pago;
    - IGNORA linhas cujo Cód não tenha exatamente 8 dígitos
      ou não comece com "10";
    - valor acumulado deve ser maior que zero;
    - cria um pedido por linha válida;
    - nota modelo 106256;
    - produto 12790;
    - quantidade 1;
    - unidade CJ;
    - local 101;
    - sem financeiro;
    - parceiro padrão 46996 enquanto não houver regra dinâmica.
    """

    SANKHYA_PEDIDOS_URL = (
        "https://api.sankhya.com.br/v1/vendas/pedidos"
    )

    NOME_ABA_XLSX = "OBRAS"

    NOTA_MODELO = 106256
    CODIGO_CLIENTE_PADRAO = 46996

    CODIGO_PRODUTO = 12790
    UNIDADE = "CJ"
    QUANTIDADE = 1

    CODIGO_LOCAL_ESTOQUE = 101
    CONTROLE = ""

    MAX_FILE_SIZE = 10 * 1024 * 1024

    CODPROJ_HEADERS = {
        "cod",
        "codigo",
        "codproj",
        "codprojeto",
        "codigoprojeto",
    }

    # Na planilha real a coluna K se chama:
    # "Valor Acumulado Pago"
    VALOR_HEADERS = {
        "valoracumuladopago",
        "valoracumulado",
        "vlracumulado",
        "valor",
    }

    def __init__(
        self,
        sankhya_client: SankhyaClient,
    ) -> None:
        self._sankhya_client = sankhya_client

    # ============================================================
    # PREVIEW
    # ============================================================

    def preview(
        self,
        *,
        filename: str,
        content: bytes,
    ) -> MaoObraPreviewResponse:
        self._validar_arquivo(
            filename=filename,
            content=content,
        )

        linhas, aba = self._extrair_linhas(
            filename=filename,
            content=content,
        )

        total_validas = sum(
            1
            for linha in linhas
            if linha.valida
        )

        total_invalidas = (
            len(linhas)
            - total_validas
        )

        valor_total_valido = Decimal("0")

        for linha in linhas:
            if (
                linha.valida
                and linha.valor_acumulado is not None
            ):
                valor_total_valido += Decimal(
                    str(linha.valor_acumulado)
                )

        return MaoObraPreviewResponse(
            arquivo=filename,
            aba=aba,
            total_linhas=len(linhas),
            total_validas=total_validas,
            total_invalidas=total_invalidas,
            valor_total_valido=float(
                self._money(valor_total_valido)
            ),
            linhas=linhas,
        )

    # ============================================================
    # PROCESSAMENTO
    # ============================================================

    async def processar(
        self,
        *,
        filename: str,
        content: bytes,
        codigo_cliente: int | None = None,
    ) -> MaoObraImportacaoResponse:
        preview = self.preview(
            filename=filename,
            content=content,
        )

        codigo_cliente_final = (
            codigo_cliente
            or self.CODIGO_CLIENTE_PADRAO
        )

        agora = datetime.now(
            ZoneInfo("America/Sao_Paulo")
        )

        data = agora.strftime("%d/%m/%Y")
        hora = agora.strftime("%H:%M")

        resultados: list[
            MaoObraResultadoLinha
        ] = []

        total_sucesso = 0
        total_erros = 0

        valor_total_processado = Decimal("0")

        for linha in preview.linhas:
            if not linha.valida:
                resultados.append(
                    MaoObraResultadoLinha(
                        linha=linha.linha,
                        codproj=linha.codproj,
                        valor_acumulado=(
                            linha.valor_acumulado
                        ),
                        sucesso=False,
                        erro=linha.erro,
                    )
                )
                continue

            if (
                linha.codproj is None
                or linha.valor_acumulado is None
            ):
                resultados.append(
                    MaoObraResultadoLinha(
                        linha=linha.linha,
                        codproj=linha.codproj,
                        valor_acumulado=(
                            linha.valor_acumulado
                        ),
                        sucesso=False,
                        erro=(
                            "Linha inválida após "
                            "normalização."
                        ),
                    )
                )
                continue

            valor = self._money(
                Decimal(
                    str(linha.valor_acumulado)
                )
            )

            try:
                response = await self._criar_pedido(
                    codproj=linha.codproj,
                    valor=valor,
                    codigo_cliente=(
                        codigo_cliente_final
                    ),
                    filename=filename,
                    linha=linha.linha,
                    data=data,
                    hora=hora,
                )

                codigo_pedido = (
                    self._extrair_codigo_pedido(
                        response
                    )
                )

                resultados.append(
                    MaoObraResultadoLinha(
                        linha=linha.linha,
                        codproj=linha.codproj,
                        valor_acumulado=float(
                            valor
                        ),
                        sucesso=True,
                        codigo_pedido=(
                            codigo_pedido
                        ),
                    )
                )

                total_sucesso += 1
                valor_total_processado += valor

            except Exception as exc:
                logger.exception(
                    "Erro ao importar mão de obra. "
                    "Arquivo=%s | Linha=%s | "
                    "CODPROJ=%s | Valor=%s",
                    filename,
                    linha.linha,
                    linha.codproj,
                    valor,
                )

                resultados.append(
                    MaoObraResultadoLinha(
                        linha=linha.linha,
                        codproj=linha.codproj,
                        valor_acumulado=float(
                            valor
                        ),
                        sucesso=False,
                        erro=str(exc),
                    )
                )

                total_erros += 1

        total_processadas = (
            total_sucesso
            + total_erros
        )

        return MaoObraImportacaoResponse(
            arquivo=filename,
            total_linhas=(
                preview.total_linhas
            ),
            total_validas=(
                preview.total_validas
            ),
            total_invalidas=(
                preview.total_invalidas
            ),
            total_processadas=(
                total_processadas
            ),
            total_sucesso=total_sucesso,
            total_erros=total_erros,
            valor_total_processado=float(
                self._money(
                    valor_total_processado
                )
            ),
            resultados=resultados,
        )

    # ============================================================
    # SANKHYA
    # ============================================================

    async def _criar_pedido(
        self,
        *,
        codproj: int,
        valor: Decimal,
        codigo_cliente: int,
        filename: str,
        linha: int,
        data: str,
        hora: str,
    ) -> dict[str, Any]:
        valor_float = float(valor)

        observacao = (
            "IMPORTACAO MAO DE OBRA"
            f" | CODPROJ {codproj}"
            f" | ARQUIVO {filename[:80]}"
            f" | LINHA {linha}"
        )

        payload: dict[str, Any] = {
            "notaModelo": self.NOTA_MODELO,
            "data": data,
            "hora": hora,
            "codigoCliente": (
                codigo_cliente
            ),
            "CODPROJ": codproj,
            "observacao": observacao,
            "valorTotal": valor_float,
            "itens": [
                {
                    "sequencia": 1,
                    "codigoProduto": (
                        self.CODIGO_PRODUTO
                    ),
                    "unidade": (
                        self.UNIDADE
                    ),
                    "quantidade": (
                        self.QUANTIDADE
                    ),
                    "controle": (
                        self.CONTROLE
                    ),
                    "codigoLocalEstoque": (
                        self.CODIGO_LOCAL_ESTOQUE
                    ),
                    "valorUnitario": (
                        valor_float
                    ),
                }
            ],
            "financeiros": [],
        }

        logger.info(
            "Criando pedido de mão de obra. "
            "CODPROJ=%s | Valor=%s | Linha=%s",
            codproj,
            valor,
            linha,
        )

        return await self._sankhya_client.post_rest(
            url=self.SANKHYA_PEDIDOS_URL,
            request_body=payload,
        )

    @staticmethod
    def _extrair_codigo_pedido(
        response: dict[str, Any],
    ) -> str | None:
        retorno = response.get(
            "retorno"
        )

        if isinstance(retorno, dict):
            codigo = retorno.get(
                "codigoPedido"
            )

            if codigo is not None:
                return str(codigo)

        codigo = response.get(
            "codigo"
        )

        if codigo is not None:
            return str(codigo)

        return None

    # ============================================================
    # ARQUIVO
    # ============================================================

    def _validar_arquivo(
        self,
        *,
        filename: str,
        content: bytes,
    ) -> None:
        if not filename:
            raise ValueError(
                "Nome do arquivo não informado."
            )

        if not content:
            raise ValueError(
                "O arquivo está vazio."
            )

        if len(content) > self.MAX_FILE_SIZE:
            raise ValueError(
                "O arquivo excede o limite "
                "de 10 MB."
            )

        suffix = (
            Path(filename)
            .suffix
            .lower()
        )

        if suffix not in {
            ".xlsx",
            ".csv",
        }:
            raise ValueError(
                "Formato não suportado. "
                "Envie XLSX ou CSV."
            )

    def _extrair_linhas(
        self,
        *,
        filename: str,
        content: bytes,
    ) -> tuple[
        list[MaoObraLinhaPreview],
        str | None,
    ]:
        suffix = (
            Path(filename)
            .suffix
            .lower()
        )

        if suffix == ".xlsx":
            return self._ler_xlsx(
                content
            )

        return (
            self._ler_csv(content),
            None,
        )

    def _ler_xlsx(
        self,
        content: bytes,
    ) -> tuple[
        list[MaoObraLinhaPreview],
        str,
    ]:
        try:
            workbook = load_workbook(
                filename=io.BytesIO(
                    content
                ),
                read_only=True,
                data_only=True,
            )

        except Exception as exc:
            raise ValueError(
                "Não foi possível ler "
                "o arquivo XLSX."
            ) from exc

        try:
            nome_aba = self.NOME_ABA_XLSX

            if nome_aba not in workbook.sheetnames:
                abas_disponiveis = ", ".join(
                    workbook.sheetnames
                )

                raise ValueError(
                    f"A aba '{nome_aba}' não foi encontrada "
                    f"na planilha. "
                    f"Abas disponíveis: "
                    f"{abas_disponiveis}"
                )

            worksheet = workbook[nome_aba]

            matriz = [
                list(row)
                for row in worksheet.iter_rows(
                    values_only=True
                )
            ]

            linhas = (
                self._extrair_da_matriz(
                    matriz
                )
            )

            return (
                linhas,
                worksheet.title,
            )

        finally:
            workbook.close()

    def _ler_csv(
        self,
        content: bytes,
    ) -> list[MaoObraLinhaPreview]:
        texto = self._decode_csv(
            content
        )

        sample = texto[:8192]

        try:
            dialect = csv.Sniffer().sniff(
                sample,
                delimiters=";,\t",
            )

            delimiter = (
                dialect.delimiter
            )

        except csv.Error:
            delimiter = ";"

        reader = csv.reader(
            io.StringIO(texto),
            delimiter=delimiter,
        )

        matriz = [
            list(row)
            for row in reader
        ]

        return self._extrair_da_matriz(
            matriz
        )

    @staticmethod
    def _decode_csv(
        content: bytes,
    ) -> str:
        for encoding in (
            "utf-8-sig",
            "utf-8",
            "latin-1",
        ):
            try:
                return content.decode(
                    encoding
                )

            except UnicodeDecodeError:
                continue

        raise ValueError(
            "Não foi possível identificar "
            "a codificação do CSV."
        )

    # ============================================================
    # EXTRAÇÃO DAS COLUNAS
    # ============================================================

    def _extrair_da_matriz(
        self,
        matriz: list[list[Any]],
    ) -> list[MaoObraLinhaPreview]:
        if not matriz:
            raise ValueError(
                "A planilha não possui dados."
            )

        (
            header_index,
            codproj_index,
            valor_index,
        ) = self._localizar_cabecalho(
            matriz
        )

        linhas: list[
            MaoObraLinhaPreview
        ] = []

        for row_index in range(
            header_index + 1,
            len(matriz),
        ):
            row = matriz[row_index]

            codproj_raw = (
                row[codproj_index]
                if codproj_index < len(row)
                else None
            )

            valor_raw = (
                row[valor_index]
                if valor_index < len(row)
                else None
            )

            # ----------------------------------------------------
            # FILTRO DE NEGÓCIO
            #
            # As linhas com códigos antigos/internos como:
            # 1, 2, 23, 36, 50...
            #
            # NÃO fazem parte da importação.
            #
            # Só entram no preview/processamento projetos com:
            # - 8 dígitos;
            # - prefixo "10".
            #
            # Ex.: 10050000
            # ----------------------------------------------------
            if not self._eh_codproj_importavel(
                codproj_raw
            ):
                continue

            linha_planilha = (
                row_index + 1
            )

            linhas.append(
                self._normalizar_linha(
                    linha=linha_planilha,
                    codproj_raw=(
                        codproj_raw
                    ),
                    valor_raw=valor_raw,
                )
            )

        if not linhas:
            raise ValueError(
                "Nenhuma linha importável foi encontrada. "
                "São aceitos somente projetos cujo Cód "
                "possua 8 dígitos e comece com 10."
            )

        return linhas

    def _localizar_cabecalho(
        self,
        matriz: list[list[Any]],
    ) -> tuple[int, int, int]:
        limite = min(
            len(matriz),
            30,
        )

        for row_index in range(
            limite
        ):
            row = matriz[row_index]

            normalized = [
                self._normalize_header(
                    value
                )
                for value in row
            ]

            codproj_index = (
                self._find_header_index(
                    normalized,
                    self.CODPROJ_HEADERS,
                )
            )

            valor_index = (
                self._find_header_index(
                    normalized,
                    self.VALOR_HEADERS,
                )
            )

            if (
                codproj_index is not None
                and valor_index is not None
            ):
                return (
                    row_index,
                    codproj_index,
                    valor_index,
                )

        raise ValueError(
            "Não encontrei as colunas "
            "'Cód' e 'Valor Acumulado Pago' "
            "nas primeiras 30 linhas."
        )

    @staticmethod
    def _find_header_index(
        headers: list[str],
        accepted: set[str],
    ) -> int | None:
        for index, header in enumerate(
            headers
        ):
            if header in accepted:
                return index

        return None

    @staticmethod
    def _normalize_header(
        value: Any,
    ) -> str:
        if value is None:
            return ""

        text = str(value).strip()

        text = unicodedata.normalize(
            "NFD",
            text,
        )

        text = "".join(
            char
            for char in text
            if unicodedata.category(
                char
            )
            != "Mn"
        )

        text = text.casefold()

        return re.sub(
            r"[^a-z0-9]",
            "",
            text,
        )

    # ============================================================
    # NORMALIZAÇÃO / VALIDAÇÃO
    # ============================================================

    @staticmethod
    def _eh_codproj_importavel(
        value: Any,
    ) -> bool:
        """
        Retorna True somente para códigos de projeto no padrão:
        10050000

        Ou seja:
        - inteiro;
        - exatamente 8 dígitos;
        - começa com 10.
        """

        if value is None:
            return False

        if isinstance(value, bool):
            return False

        text = str(value).strip()

        if not text:
            return False

        try:
            numero = Decimal(text)

        except (
            InvalidOperation,
            ValueError,
        ):
            return False

        if (
            numero
            != numero.to_integral_value()
        ):
            return False

        codigo = str(
            int(numero)
        )

        return (
            len(codigo) == 8
            and codigo.startswith("10")
        )

    def _normalizar_linha(
        self,
        *,
        linha: int,
        codproj_raw: Any,
        valor_raw: Any,
    ) -> MaoObraLinhaPreview:
        erros: list[str] = []

        codproj: int | None = None
        valor: Decimal | None = None

        try:
            codproj = self._parse_codproj(
                codproj_raw
            )

        except ValueError as exc:
            erros.append(
                str(exc)
            )

        try:
            valor = self._parse_money(
                valor_raw
            )

        except ValueError as exc:
            erros.append(
                str(exc)
            )

        if (
            valor is not None
            and valor <= 0
        ):
            erros.append(
                "Valor acumulado deve "
                "ser maior que zero."
            )

        return MaoObraLinhaPreview(
            linha=linha,
            codproj=codproj,
            valor_acumulado=(
                float(
                    self._money(valor)
                )
                if valor is not None
                else None
            ),
            valida=not erros,
            erro=(
                " | ".join(erros)
                if erros
                else None
            ),
        )

    @staticmethod
    def _parse_codproj(
        value: Any,
    ) -> int:
        if value is None:
            raise ValueError(
                "Cód não informado."
            )

        if isinstance(value, bool):
            raise ValueError(
                "Cód inválido."
            )

        text = str(value).strip()

        if not text:
            raise ValueError(
                "Cód não informado."
            )

        try:
            numero = Decimal(text)

        except (
            InvalidOperation,
            ValueError,
        ) as exc:
            raise ValueError(
                f"Cód inválido: {value}"
            ) from exc

        if (
            numero
            != numero.to_integral_value()
        ):
            raise ValueError(
                f"Cód inválido: {value}"
            )

        codproj = int(numero)
        codigo = str(codproj)

        if len(codigo) != 8:
            raise ValueError(
                f"Cód {codproj} inválido: "
                "o projeto deve possuir "
                "exatamente 8 dígitos."
            )

        if not codigo.startswith("10"):
            raise ValueError(
                f"Cód {codproj} inválido: "
                "o projeto deve iniciar com 10."
            )

        return codproj

    @staticmethod
    def _parse_money(
        value: Any,
    ) -> Decimal:
        if value is None:
            raise ValueError(
                "Valor acumulado não informado."
            )

        if isinstance(value, Decimal):
            return value

        if isinstance(
            value,
            (int, float),
        ):
            return Decimal(
                str(value)
            )

        text = str(value).strip()

        if not text:
            raise ValueError(
                "Valor acumulado não informado."
            )

        text = (
            text
            .replace("R$", "")
            .replace("\xa0", "")
            .replace(" ", "")
        )

        text = re.sub(
            r"[^0-9,.\-]",
            "",
            text,
        )

        if (
            "," in text
            and "." in text
        ):
            if (
                text.rfind(",")
                > text.rfind(".")
            ):
                # Ex.: 16.456,60
                text = (
                    text
                    .replace(".", "")
                    .replace(",", ".")
                )

            else:
                # Ex.: 16,456.60
                text = (
                    text
                    .replace(",", "")
                )

        elif "," in text:
            # Ex.: 16456,60
            text = text.replace(
                ",",
                ".",
            )

        try:
            return Decimal(text)

        except InvalidOperation as exc:
            raise ValueError(
                f"Valor acumulado inválido: "
                f"{value}"
            ) from exc

    @staticmethod
    def _money(
        value: Decimal,
    ) -> Decimal:
        return value.quantize(
            Decimal("0.01"),
            rounding=ROUND_HALF_UP,
        )


_importacao_mao_obra_service: (
    ImportacaoMaoObraService | None
) = None


def get_importacao_mao_obra_service(
    sankhya_client: SankhyaClient = Depends(
        get_sankhya_client
    ),
) -> ImportacaoMaoObraService:
    global _importacao_mao_obra_service

    if (
        _importacao_mao_obra_service
        is None
    ):
        _importacao_mao_obra_service = (
            ImportacaoMaoObraService(
                sankhya_client=sankhya_client
            )
        )

    return _importacao_mao_obra_service